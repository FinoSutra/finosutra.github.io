# -*- coding: utf-8 -*-
"""Finosutra IND AS 116 Lease Template v3.0 — builder.

Builds the workbook that downloadTemplate() in app.js serves. It is generated here
rather than in the browser because the dropdown lists, the per-cell help prompts and
the header notes cannot be written by xlsx-js-style.

    python scripts/build_lease_template.py

The dropdown VALUES are load-bearing: processUploadedLeases() in app.js maps them by
substring ('percent' -> pct, 'fixed amount' -> amt, 'index' -> cpi, 'beg'/'adv' ->
advance). Changing the wording without changing that mapping will silently
mis-classify leases.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ── Palette ──────────────────────────────────────────────────────────────
BRAND   = "1E1B4B"   # deep indigo
BRAND2  = "4F46E5"   # indigo  (optional headers)
AMBER   = "B45309"   # amber   (required headers)
WHITE   = "FFFFFF"
GREY_BG = "F3F4F6"
ALT_BG  = "FAFAFA"
BORDER  = "D1D5DB"
STOPBG  = "FEE2E2"
STOPFG  = "991B1B"
NOTEBG  = "EEF2FF"
OKBG    = "ECFDF5"

F = "Arial"
thin = Side(style="thin", color=BORDER)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ═════════════════════════════════════════════════════════════════════════
# Column definition — order here IS the sheet order
# key: header text (must stay parser-compatible), required, width,
#      number_format, tooltip title, tooltip body, common mistake
# ═════════════════════════════════════════════════════════════════════════
COLUMNS = [
    dict(h="Lease Name *", req=True, w=30, fmt="@",
         tt="Lease Name (required)",
         tip="Unique name for this lease. Re-uploading the same name UPDATES that "
             "lease instead of creating a duplicate.\nExample: Mumbai Office - Floor 4",
         miss="Two leases with the same name - the second overwrites the first."),
    dict(h="Lessor Name", req=False, w=22, fmt="@",
         tt="Lessor Name (optional)",
         tip="The landlord / owner of the asset.\nExample: Raheja Corp",
         miss="-"),
    dict(h="Entity / Lessee", req=False, w=22, fmt="@",
         tt="Entity / Lessee (optional)",
         tip="Your legal entity that signed the lease.\nExample: ABC Pvt. Ltd.",
         miss="-"),
    dict(h="Asset Category", req=False, w=18, fmt="@",
         tt="Asset Category (optional)",
         tip="Pick from the dropdown. Used to group leases in your disclosure note. "
             "If nothing fits, type your own.",
         miss="-"),
    dict(h="Start Date *", req=True, w=15, fmt="@",
         tt="Start Date (required)",
         tip="Lease COMMENCEMENT date - the date you actually got the asset, NOT the "
             "date the agreement was signed.\nType as DD/MM/YYYY, e.g. 01/03/2026\n"
             "This column is text-formatted so Excel will not reformat your date.",
         miss="Entering the signing date, or letting Excel flip it to MM/DD/YYYY."),
    dict(h="Term (months) *", req=True, w=15, fmt="0",
         tt="Term in months (required)",
         tip="Total lease term in MONTHS - including any rent-free months.\n"
             "3 years = 36     5 years = 60     10 years = 120\nWhole numbers only.",
         miss="Entering years (5) instead of months (60)."),
    dict(h="Rent / Period (₹) *", req=True, w=18, fmt="#,##0.00",
         tt="Rent per period (required)",
         tip="Rent for ONE payment period - not per year.\nIf Frequency = Quarterly, "
             "enter the QUARTERLY amount.\nEnter the STARTING rent only; increases are "
             "handled in columns K-N.\nNumbers only - no commas, no Rs sign.",
         miss="Entering annual rent when Frequency is Monthly."),
    dict(h="Frequency", req=False, w=15, fmt="@",
         tt="Payment Frequency (optional)",
         tip="How often rent is paid. Pick from the dropdown.\nBlank = Monthly.",
         miss="-"),
    dict(h="IBR (% p.a.) *", req=True, w=14, fmt="0.00",
         tt="Incremental Borrowing Rate (required)",
         tip="Your incremental borrowing rate, % per YEAR.\nEnter 10 for 10%, "
             "10.5 for 10.5%.\nTypical Indian IBR is 8% to 14%.",
         miss="Entering 0.10 instead of 10."),
    dict(h="Payment Timing", req=False, w=26, fmt="@",
         tt="Payment Timing (optional)",
         tip="WHEN within each period the rent is paid.\n"
             "End of period (Arrears) = paid on the last day - most common in India.\n"
             "Beginning of period (Advance) = paid on the first day, in advance.\n"
             "Blank = End of period.",
         miss="Advance rent booked as arrears - overstates the liability."),
    dict(h="Escalation Type", req=False, w=26, fmt="@",
         tt="Escalation Type (optional)",
         tip="Does the rent increase during the lease? Pick from the dropdown.\n"
             "None = same rent throughout\nPercent = fixed % increase (fill column L)\n"
             "Fixed amount = fixed Rs increase (fill column M)\n"
             "CPI = index-linked (fill column L)",
         miss="Choosing an escalation type but leaving L/M/N blank."),
    dict(h="Escalation % p.a.", req=False, w=17, fmt="0.00",
         tt="Escalation percentage",
         tip="ONLY if Escalation Type = Percent or CPI.\nEnter 5 for a 5% increase at "
             "each step.\nLeave blank otherwise.",
         miss="Entering 0.05 instead of 5."),
    dict(h="Escalation ₹ per step", req=False, w=20, fmt="#,##0.00",
         tt="Escalation rupees per step",
         tip="ONLY if Escalation Type = Fixed amount.\nThe rupee increase added at each "
             "step.\nExample: 5000 means rent rises by Rs 5,000 each step.\n"
             "Leave blank otherwise.",
         miss="-"),
    dict(h="Escalation Interval (yrs)", req=False, w=21, fmt="0",
         tt="Escalation Interval - READ THIS",
         tip="How often the rent steps up.\n"
             "1 = increases EVERY YEAR  (most common)\n2 = every 2 years\n"
             "3 = every 3 years\n"
             "Do NOT leave this blank and do NOT enter 0 when using escalation.",
         miss="Leaving blank or entering 0 - escalation then does not apply yearly."),
    dict(h="Rent-Free Months", req=False, w=17, fmt="0",
         tt="Rent-Free Months (optional)",
         tip="Number of rent-free months at the START of the lease.\n"
             "These months STILL count inside the term in column F, and interest "
             "still accrues on them.\nEnter 0 if none.",
         miss="Excluding rent-free months from the term in column F."),
    dict(h="IDC (₹)", req=False, w=15, fmt="#,##0.00",
         tt="Initial Direct Costs (optional)",
         tip="Brokerage, legal fees, stamp duty paid to obtain the lease.\n"
             "ADDED to the ROU asset.\nEnter 0 if none.",
         miss="-"),
    dict(h="Incentive (₹)", req=False, w=15, fmt="#,##0.00",
         tt="Lease Incentive (optional)",
         tip="Incentives RECEIVED from the lessor - fit-out contribution, cash back.\n"
             "DEDUCTED from the ROU asset.\nEnter 0 if none.",
         miss="Entering it as a negative number - enter the positive amount."),
    dict(h="Restoration Cost (₹)", req=False, w=20, fmt="#,##0.00",
         tt="Restoration / Make-good (optional)",
         tip="Estimated cost to restore the asset at the end of the lease.\n"
             "ADDED to the ROU asset.\nEnter 0 if none.",
         miss="-"),
    dict(h="Short-term / Low-value", req=False, w=22, fmt="@",
         tt="IND AS 116 exemption (optional)",
         tip="Leave BLANK for normal leases.\n"
             "short-term = lease term is 12 months or less\n"
             "low-value  = asset is low value when new\n"
             "If set: NO ROU asset and NO lease liability. Rent is expensed "
             "straight-line instead.",
         miss="Using it for a 5-year lease - the exemption then wrongly removes the asset."),
]

NCOL       = len(COLUMNS)
LAST       = get_column_letter(NCOL)
HDR_ROW    = 4
SAMPLE_1   = 5
SAMPLE_2   = 6
DATA_START = 7
MAX_LEASES = 20
DATA_END   = DATA_START + MAX_LEASES - 1     # 26
STOP_ROW   = DATA_END + 1                    # 27

# Dropdown lists — values chosen so the import parser maps them correctly
DD_FREQ   = '"Monthly,Quarterly,Half-Yearly,Annual"'
DD_TIMING = '"End of period (Arrears),Beginning of period (Advance)"'
DD_ESC    = '"None (flat rent),Percent (% increase),Fixed amount (Rs step-up),CPI / Index-linked"'
DD_EXEMPT = '"short-term,low-value"'
CATEGORIES = ["Office Space", "Warehouse", "Retail Store", "Plant & Machinery",
              "Vehicle", "IT Equipment", "Equipment", "Land", "Building", "Other"]

# ═════════════════════════════════════════════════════════════════════════
# SHEET 1 — START HERE
# ═════════════════════════════════════════════════════════════════════════
s1 = wb.active
s1.title = "START HERE"
s1.sheet_view.showGridLines = False
s1.column_dimensions["A"].width = 4
s1.column_dimensions["B"].width = 104

def s1row(r, text, *, size=11, bold=False, color="111827", fill=None,
          height=None, italic=False):
    c = s1.cell(row=r, column=2, value=text)
    c.font = Font(name=F, size=size, bold=bold, color=color, italic=italic)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
        s1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=fill)
    if height:
        s1.row_dimensions[r].height = height
    return c

s1row(1, "Finosutra  —  IND AS 116 Lease Template", size=20, bold=True,
      color=BRAND, height=34)
s1row(2, "Version 3.0   |   IND AS 116 / IFRS 16   |   Read this page once, "
         "then go to the 'Lease Register' tab.", size=10, color="6B7280", height=20)
s1.row_dimensions[3].height = 8

s1row(4, "How to fill this template", size=14, bold=True, color=BRAND, height=26)

steps = [
    ("Step 1", "Open the 'Lease Register' tab. Your data goes in rows 7 to 26 — "
               "that is 20 leases per file."),
    ("Step 2", "Rows 5 and 6 are worked SAMPLES, shown in grey italics. "
               "Delete both rows before you upload."),
    ("Step 3", "Click any cell to see a yellow help box telling you exactly what to "
               "enter. Hover over any orange or blue heading for a fuller note."),
    ("Step 4", "Orange headings are REQUIRED: Lease Name, Start Date, Term, "
               "Rent per Period, IBR. Blue headings are optional — leave blank to "
               "use the default."),
    ("Step 5", "Columns with a small arrow on the right are dropdowns. Always pick "
               "from the list rather than typing, so the upload reads them correctly."),
    ("Step 6", "Save the file, then upload it in Finosutra → Lease Workspace → "
               "Upload Portfolio. Check the preview screen before you confirm."),
]
r = 5
for label, body in steps:
    s1.row_dimensions[r].height = 32
    a = s1.cell(row=r, column=1, value=label.replace("Step ", ""))
    a.font = Font(name=F, size=11, bold=True, color=WHITE)
    a.fill = PatternFill("solid", fgColor=BRAND2)
    a.alignment = Alignment(horizontal="center", vertical="center")
    c = s1.cell(row=r, column=2, value=body)
    c.font = Font(name=F, size=11, color="111827")
    c.alignment = Alignment(vertical="center", wrap_text=True)
    r += 1

r += 1
s1row(r, "The five things people get wrong", size=14, bold=True, color=BRAND,
      height=28); r += 1
pitfalls = [
    "Term is in MONTHS, not years. A 5-year lease is 60.",
    "Rent is per PAYMENT PERIOD, not per year. Quarterly frequency means the "
    "quarterly amount.",
    "IBR is a whole percentage. Enter 10 for 10% — not 0.10.",
    "If rent increases every year, set 'Escalation Interval (yrs)' to 1. "
    "Leaving it blank does not mean yearly.",
    "Start Date is the COMMENCEMENT date, typed as DD/MM/YYYY.",
]
for p in pitfalls:
    s1.row_dimensions[r].height = 30
    a = s1.cell(row=r, column=1, value="!")
    a.font = Font(name=F, size=12, bold=True, color=STOPFG)
    a.alignment = Alignment(horizontal="center", vertical="center")
    c = s1.cell(row=r, column=2, value=p)
    c.font = Font(name=F, size=11, color="111827")
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor=NOTEBG)
    r += 1

r += 1
s1.row_dimensions[r].height = 46
c = s1.cell(row=r, column=2,
            value="LIMIT — 20 leases per file.\nThis keeps the upload fast and lets you "
                  "actually check the preview screen. If you have more than 20 leases, "
                  "save a second copy of this file and upload it separately.")
c.font = Font(name=F, size=11, bold=True, color=STOPFG)
c.fill = PatternFill("solid", fgColor=STOPBG)
c.alignment = Alignment(vertical="center", wrap_text=True)
s1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=STOPBG)
r += 2

s1row(r, "Tabs in this workbook", size=14, bold=True, color=BRAND, height=26); r += 1
for t, d in [("Lease Register", "Where you enter your leases. This is the only tab we read."),
             ("Field Guide", "Every column explained, with the common mistake for each."),
             ("Examples", "Five fully worked leases, including a yearly-escalating one.")]:
    s1.row_dimensions[r].height = 20
    c = s1.cell(row=r, column=2, value="%-18s  %s" % (t, d))
    c.font = Font(name=F, size=11, color="374151")
    c.alignment = Alignment(vertical="center")
    r += 1

# ═════════════════════════════════════════════════════════════════════════
# SHEET 2 — Lease Register
# ═════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Lease Register")
ws.sheet_view.showGridLines = False

# Row 1 — title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
t = ws.cell(row=1, column=1, value="Finosutra   |   IND AS 116 Lease Register")
t.font = Font(name=F, size=16, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=BRAND)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

# Row 2 — subtitle
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
t = ws.cell(row=2, column=2 - 1,
            value="Version 3.0   |   Maximum 20 leases per file   |   "
                  "Enter your data in rows 7-26   |   Click any cell for help")
t.font = Font(name=F, size=10, color="C7D2FE")
t.fill = PatternFill("solid", fgColor=BRAND)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 18

# Row 3 — legend + live counter
ws.row_dimensions[3].height = 20
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=5)
lg = ws.cell(row=3, column=1, value="ORANGE heading = required     BLUE heading = optional")
lg.font = Font(name=F, size=9, bold=True, color="374151")
lg.fill = PatternFill("solid", fgColor=GREY_BG)
lg.alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=13)
lg2 = ws.cell(row=3, column=6,
              value="Rows 5-6 are samples - delete them before uploading.")
lg2.font = Font(name=F, size=9, italic=True, color="6B7280")
lg2.fill = PatternFill("solid", fgColor=GREY_BG)
lg2.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells(start_row=3, start_column=14, end_row=3, end_column=NCOL)
cnt = ws.cell(row=3, column=14,
              value="Enter up to %d leases  ->  rows %d to %d"
                    % (MAX_LEASES, DATA_START, DATA_END))
cnt.font = Font(name=F, size=9, bold=True, color=BRAND2)
cnt.fill = PatternFill("solid", fgColor=GREY_BG)
cnt.alignment = Alignment(horizontal="right", vertical="center", indent=1)

# Row 4 — headers
ws.row_dimensions[HDR_ROW].height = 42
for i, col in enumerate(COLUMNS, start=1):
    c = ws.cell(row=HDR_ROW, column=i, value=col["h"])
    c.font = Font(name=F, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=AMBER if col["req"] else BRAND2)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    note = col["tip"]
    if col["miss"] != "-":
        note += "\n\nCommon mistake: " + col["miss"]
    cm = Comment(note, "Finosutra")
    cm.width, cm.height = 300, 190
    c.comment = cm
    ws.column_dimensions[get_column_letter(i)].width = col["w"]

# Rows 5-6 — samples
samples = [
    ["SAMPLE 1 - delete this row", "Raheja Corp", "ABC Pvt. Ltd.", "Office Space",
     "01/04/2025", 36, 50000, "Monthly", 10.5, "End of period (Arrears)",
     "None (flat rent)", None, None, None, 0, 0, 0, 0, None],
    ["SAMPLE 2 - delete this row", "Industrial Parks Ltd", "ABC Pvt. Ltd.", "Warehouse",
     "01/07/2025", 60, 95000, "Monthly", 10, "End of period (Arrears)",
     "Percent (% increase)", 5, None, 1, 2, 10000, 0, 0, None],
]
for ri, row in enumerate(samples):
    rr = SAMPLE_1 + ri
    ws.row_dimensions[rr].height = 20
    for ci, v in enumerate(row, start=1):
        c = ws.cell(row=rr, column=ci, value=v)
        c.font = Font(name=F, size=10, italic=True, color="6B7280")
        c.fill = PatternFill("solid", fgColor=GREY_BG)
        c.border = box
        c.number_format = COLUMNS[ci - 1]["fmt"]
        c.alignment = Alignment(horizontal="left" if COLUMNS[ci - 1]["fmt"] == "@"
                                else "right", vertical="center")

# Rows 7-26 — 20 blank, formatted input rows
for rr in range(DATA_START, DATA_END + 1):
    ws.row_dimensions[rr].height = 19
    band = ALT_BG if (rr - DATA_START) % 2 else WHITE
    for ci, col in enumerate(COLUMNS, start=1):
        c = ws.cell(row=rr, column=ci)
        c.font = Font(name=F, size=10, color="111827")
        c.fill = PatternFill("solid", fgColor=band)
        c.border = box
        c.number_format = col["fmt"]
        c.alignment = Alignment(horizontal="left" if col["fmt"] == "@" else "right",
                                vertical="center")

# Row 27 — hard stop banner
ws.row_dimensions[STOP_ROW].height = 34
ws.merge_cells(start_row=STOP_ROW, start_column=1, end_row=STOP_ROW, end_column=NCOL)
sc = ws.cell(row=STOP_ROW, column=1,
             value="STOP  —  Maximum 20 leases per file. Do not add rows below this line. "
                   "If you have more than 20 leases, save a second copy of this file and "
                   "upload it separately.")
sc.font = Font(name=F, size=11, bold=True, color=STOPFG)
sc.fill = PatternFill("solid", fgColor=STOPBG)
sc.alignment = Alignment(horizontal="center", vertical="center")

# ── Data validation: dropdowns + click-tooltips on every column ──────────
DV_RANGE_END = 200   # keep dropdowns alive even if a user adds rows

def add_dv(col_idx, formula1=None, allow_blank=True):
    col = COLUMNS[col_idx - 1]
    letter = get_column_letter(col_idx)
    if formula1:
        dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank,
                            showDropDown=False, showErrorMessage=True,
                            errorTitle="Please pick from the list",
                            error="Use the dropdown arrow on the right of this cell. "
                                  "Typing your own text here can make the upload "
                                  "misread this lease.",
                            errorStyle="warning")
    else:
        dv = DataValidation(allow_blank=True, showErrorMessage=False)
    dv.promptTitle = col["tt"][:32]
    dv.prompt = col["tip"][:255]
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add("%s%d:%s%d" % (letter, SAMPLE_1, letter, DV_RANGE_END))

DROPDOWNS = {
    4:  "Lists!$A$2:$A$11",   # Asset Category
    8:  DD_FREQ,              # Frequency
    10: DD_TIMING,            # Payment Timing
    11: DD_ESC,               # Escalation Type
    19: DD_EXEMPT,            # Short-term / Low-value
}
for i in range(1, NCOL + 1):
    add_dv(i, DROPDOWNS.get(i))

ws.freeze_panes = "A5"
ws.sheet_properties.tabColor = BRAND2

# ═════════════════════════════════════════════════════════════════════════
# SHEET 3 — Field Guide
# ═════════════════════════════════════════════════════════════════════════
fg = wb.create_sheet("Field Guide")
fg.sheet_view.showGridLines = False
widths = [7, 24, 11, 54, 34, 40]
for i, w in enumerate(widths, start=1):
    fg.column_dimensions[get_column_letter(i)].width = w

fg.merge_cells("A1:F1")
c = fg.cell(row=1, column=1, value="Field Guide  —  every column explained")
c.font = Font(name=F, size=15, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=BRAND)
c.alignment = Alignment(vertical="center", indent=1)
fg.row_dimensions[1].height = 30

hdrs = ["Col", "Field", "Required?", "What to enter", "Valid values / example",
        "Common mistake"]
fg.row_dimensions[2].height = 24
for i, h in enumerate(hdrs, start=1):
    c = fg.cell(row=2, column=i, value=h)
    c.font = Font(name=F, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND2)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = box

VALID = {
    1: "Any text. Must be unique.",
    2: "Any text.",
    3: "Any text.",
    4: "Dropdown: Office Space, Warehouse, Retail Store, Plant & Machinery, "
       "Vehicle, IT Equipment, Equipment, Land, Building, Other",
    5: "DD/MM/YYYY   e.g. 01/03/2026",
    6: "Whole number.  60 = five years",
    7: "Number only.  95000",
    8: "Dropdown: Monthly / Quarterly / Half-Yearly / Annual",
    9: "Number.  10 = 10%,  10.5 = 10.5%",
    10: "Dropdown: End of period (Arrears) / Beginning of period (Advance)",
    11: "Dropdown: None / Percent / Fixed amount / CPI",
    12: "Number.  5 = 5% per step",
    13: "Number.  5000 = Rs 5,000 per step",
    14: "1 = yearly,  2 = every 2 yrs,  3 = every 3 yrs",
    15: "Whole number.  0 if none",
    16: "Number.  0 if none",
    17: "Number.  0 if none",
    18: "Number.  0 if none",
    19: "Blank, or: short-term / low-value",
}
r = 3
for i, col in enumerate(COLUMNS, start=1):
    fg.row_dimensions[r].height = 42
    vals = [get_column_letter(i), col["h"].replace(" *", ""),
            "REQUIRED" if col["req"] else "Optional",
            col["tip"].replace("\n", "  "), VALID[i], col["miss"]]
    for ci, v in enumerate(vals, start=1):
        c = fg.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=9,
                      bold=(ci == 3 and col["req"]),
                      color=AMBER if (ci == 3 and col["req"]) else "111827")
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if ci in (1, 3) else "left")
        c.border = box
        c.fill = PatternFill("solid", fgColor=WHITE if r % 2 else ALT_BG)
    r += 1

r += 1
fg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = fg.cell(row=r, column=1, value="Notes")
c.font = Font(name=F, size=12, bold=True, color=BRAND)
r += 1
notes = [
    "Rent-free months are part of the lease term. A 60-month lease with 2 rent-free "
    "months is still 60 in column F — interest accrues during those 2 months.",
    "IDC and Restoration Cost are ADDED to the ROU asset. Incentive is DEDUCTED. "
    "Enter all three as positive numbers.",
    "Short-term / low-value leases produce NO ROU asset and NO lease liability. "
    "The rent is expensed on a straight-line basis instead.",
    "Re-uploading a lease with a name that already exists UPDATES it rather than "
    "creating a duplicate. Keep your lease names stable between uploads.",
    "Do not rename, reorder, or delete the column headings in row 4 of the Lease "
    "Register — the upload matches on those exact headings.",
]
for n in notes:
    fg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    fg.row_dimensions[r].height = 30
    c = fg.cell(row=r, column=1, value="•  " + n)
    c.font = Font(name=F, size=10, color="374151")
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    r += 1

# ═════════════════════════════════════════════════════════════════════════
# SHEET 4 — Examples
# ═════════════════════════════════════════════════════════════════════════
ex = wb.create_sheet("Examples")
ex.sheet_view.showGridLines = False
ex.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
c = ex.cell(row=1, column=1, value="Worked examples  —  copy the pattern that matches "
                                   "your lease")
c.font = Font(name=F, size=15, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=BRAND)
c.alignment = Alignment(vertical="center", indent=1)
ex.row_dimensions[1].height = 30

EXAMPLES = [
    ("A. Simple flat-rent office lease — no increases, paid at month end",
     ["Bengaluru Office", "Prestige Estates", "ABC Pvt. Ltd.", "Office Space",
      "01/04/2025", 36, 50000, "Monthly", 10.5, "End of period (Arrears)",
      "None (flat rent)", None, None, None, 0, 0, 0, 0, None]),
    ("B. Rent rising 5% EVERY YEAR — note Escalation Interval = 1",
     ["Cuttack Branch", "Sahoo Properties", "XYZ Ltd.", "Office Space",
      "01/03/2026", 60, 95000, "Monthly", 10, "End of period (Arrears)",
      "Percent (% increase)", 5, None, 1, 0, 0, 0, 0, None]),
    ("C. Rent rising by a fixed Rs 5,000 every 3 years",
     ["Chennai Depot", "TN Logistics", "XYZ Ltd.", "Warehouse",
      "01/04/2025", 108, 60000, "Monthly", 11, "End of period (Arrears)",
      "Fixed amount (Rs step-up)", None, 5000, 3, 0, 0, 0, 0, None]),
    ("D. Quarterly rent paid IN ADVANCE, with 3 rent-free months and brokerage",
     ["Gurugram Retail", "DLF Ltd", "XYZ Ltd.", "Retail Store",
      "01/07/2025", 60, 300000, "Quarterly", 9.5, "Beginning of period (Advance)",
      "None (flat rent)", None, None, None, 3, 150000, 50000, 100000, None]),
    ("E. Short-term lease (12 months) — exemption applied, no ROU asset",
     ["Site Cabin - Project A", "Local Owner", "XYZ Ltd.", "Other",
      "01/04/2025", 12, 25000, "Monthly", 10, "End of period (Arrears)",
      "None (flat rent)", None, None, None, 0, 0, 0, 0, "short-term"]),
]

r = 3
for i, col in enumerate(COLUMNS, start=1):
    c = ex.cell(row=r, column=i, value=col["h"])
    c.font = Font(name=F, size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=AMBER if col["req"] else BRAND2)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    ex.column_dimensions[get_column_letter(i)].width = col["w"]
ex.row_dimensions[r].height = 40
r += 1

for title, row in EXAMPLES:
    ex.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    c = ex.cell(row=r, column=1, value=title)
    c.font = Font(name=F, size=10, bold=True, color=BRAND)
    c.fill = PatternFill("solid", fgColor=OKBG)
    c.alignment = Alignment(vertical="center", indent=1)
    ex.row_dimensions[r].height = 22
    r += 1
    for ci, v in enumerate(row, start=1):
        c = ex.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=10, color="111827")
        c.number_format = COLUMNS[ci - 1]["fmt"]
        c.border = box
        c.alignment = Alignment(horizontal="left" if COLUMNS[ci - 1]["fmt"] == "@"
                                else "right", vertical="center")
    ex.row_dimensions[r].height = 19
    r += 2

ex.freeze_panes = "A4"

# ═════════════════════════════════════════════════════════════════════════
# SHEET 5 — Lists (hidden, drives the Asset Category dropdown)
# ═════════════════════════════════════════════════════════════════════════
lst = wb.create_sheet("Lists")
lst.cell(row=1, column=1, value="Asset Category").font = Font(name=F, bold=True)
for i, v in enumerate(CATEGORIES, start=2):
    lst.cell(row=i, column=1, value=v).font = Font(name=F, size=10)
lst.column_dimensions["A"].width = 24
lst.sheet_state = "hidden"

wb.active = 0
# Written to the repo root, where downloadTemplate() in app.js serves it from.
# Keep LEASE_TEMPLATE_URL in app.js in step with this filename.
import os
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "Finosutra_Lease_Template_v3.xlsx")
wb.save(OUT)
print("saved:", OUT)
