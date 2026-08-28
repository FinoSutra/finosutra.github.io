# -*- coding: utf-8 -*-
"""Finosutra IND AS 109 Security Deposit Portfolio Template v1.0 — builder.

Mirrors scripts/build_lease_template.py column-for-column in structure and
convention. Builds the workbook that downloadTemplate() in sd-portfolio.html
serves. It is generated here rather than in the browser because the dropdown
lists, the per-cell help prompts and the header notes cannot be written by
xlsx-js-style.

    python scripts/build_sd_template.py

The header text is load-bearing: processUploadedDeposits() in sd-portfolio.html
strips the trailing " *" from required headers before matching them (so
"Deposit Name *" is read as "Deposit Name"). Changing a header's wording
without changing that lookup will silently drop the column.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ── Palette (identical to the Lease Template) ───────────────────────────────
BRAND   = "1E1B4B"   # deep indigo
BRAND2  = "4F46E5"   # indigo  (optional headers)
AMBER   = "B45309"   # amber   (required headers)
WHITE   = "FFFFFF"
GREY_BG = "F3F4F6"
ALT_BG  = "FAFAFA"
BORDER  = "D1D5DB"
STOPBG  = "FEE2E2"
STOPFG  = "991B1B"
NOTEBG  = "EEF2FF"
OKBG    = "ECFDF5"

F = "Arial"
thin = Side(style="thin", color=BORDER)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

# ═════════════════════════════════════════════════════════════════════════
# Column definition — order here IS the sheet order
# ═════════════════════════════════════════════════════════════════════════
COLUMNS = [
    dict(h="Deposit Name *", req=True, w=30, fmt="@",
         tt="Deposit Name (required)",
         tip="Unique name for this deposit. Re-uploading the same name UPDATES "
             "that deposit instead of creating a duplicate.\n"
             "Example: Mumbai Office Deposit",
         miss="Two deposits with the same name - the second overwrites the first."),
    dict(h="Counterparty Name", req=False, w=24, fmt="@",
         tt="Counterparty Name (optional)",
         tip="The other party to the deposit - the landlord if you paid it, or "
             "the tenant if you received it.\nExample: Raheja Corp",
         miss="-"),
    dict(h="Entity / Lessee", req=False, w=22, fmt="@",
         tt="Entity / Lessee (optional)",
         tip="Your legal entity that paid or received this deposit.\n"
             "Example: ABC Pvt. Ltd.",
         miss="-"),
    dict(h="Asset / Property Category", req=False, w=26, fmt="@",
         tt="Asset / Property Category",
         tip="Pick from the dropdown. Used to group deposits in your disclosure "
             "note. If nothing fits, type your own.",
         miss="-"),
    dict(h="Deposit Amount (₹) *", req=True, w=20, fmt="#,##0.00",
         tt="Deposit Amount (required)",
         tip="The NOMINAL amount actually paid or received - NOT its present "
             "value. The fair value and Day-1 discount are calculated for you.\n"
             "Numbers only - no commas, no Rs sign.",
         miss="Entering the present value instead of the cash amount actually paid."),
    dict(h="Market Rate (% p.a.) *", req=True, w=18, fmt="0.00",
         tt="Market Rate / EIR (required)",
         tip="The market rate of interest for a similar instrument (or your "
             "IBR), used to discount the deposit to fair value. % per YEAR.\n"
             "Enter 10 for 10%, 10.5 for 10.5%.",
         miss="Entering 0.10 instead of 10."),
    dict(h="Tenure (Years) *", req=True, w=15, fmt="0.00",
         tt="Tenure in years (required)",
         tip="Total deposit term in YEARS. Decimals are allowed for part years.\n"
             "3 years 6 months = 3.5     5 years = 5",
         miss="Entering months (42) instead of years (3.5)."),
    dict(h="Commencement Date", req=False, w=20, fmt="@",
         tt="Commencement Date (optional)",
         tip="The date the deposit was actually paid or received - NOT the "
             "lease start date if they differ.\nType as DD/MM/YYYY, e.g. "
             "01/04/2025.\nThis column is text-formatted so Excel will not "
             "reformat your date.\nLeave blank if you only need lifetime totals.",
         miss="Entering the lease start date instead of the date the deposit "
              "itself changed hands."),
    dict(h="Deposit Type", req=False, w=16, fmt="@",
         tt="Deposit Type (optional)",
         tip="Given = your entity paid the deposit (it is your financial "
             "asset).\nReceived = your entity received the deposit (it is "
             "your financial liability).\nBlank = Given.",
         miss="Leaving it blank when you actually RECEIVED the deposit - it "
              "silently defaults to Given."),
    dict(h="Reporting Standard", req=False, w=18, fmt="@",
         tt="Reporting Standard (optional)",
         tip="Which standard's account labels to use in the journal entries "
             "and notes.\nBlank = IND AS.",
         miss="-"),
]

NCOL       = len(COLUMNS)
HDR_ROW    = 4
SAMPLE_1   = 5
SAMPLE_2   = 6
DATA_START = 7
MAX_DEPOSITS = 20
DATA_END   = DATA_START + MAX_DEPOSITS - 1     # 26
STOP_ROW   = DATA_END + 1                      # 27

DD_TYPE = '"Given,Received"'
DD_STD  = '"IND AS,IFRS"'
CATEGORIES = ["Office Space", "Warehouse", "Retail Store", "Plant & Machinery",
              "Vehicle", "IT Equipment", "Equipment", "Land", "Building", "Other"]

# ═════════════════════════════════════════════════════════════════════════
# SHEET 1 — START HERE
# ═════════════════════════════════════════════════════════════════════════
s1 = wb.active
s1.title = "START HERE"
s1.sheet_view.showGridLines = False
s1.column_dimensions["A"].width = 4
s1.column_dimensions["B"].width = 104

def s1row(r, text, *, size=11, bold=False, color="111827", fill=None,
          height=None, italic=False):
    c = s1.cell(row=r, column=2, value=text)
    c.font = Font(name=F, size=size, bold=bold, color=color, italic=italic)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
        s1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=fill)
    if height:
        s1.row_dimensions[r].height = height
    return c

s1row(1, "Finosutra  —  IND AS 109 Security Deposit Template", size=20, bold=True,
      color=BRAND, height=34)
s1row(2, "Version 1.0   |   IND AS 109 / IFRS 9 (with IND AS 116 / IFRS 16 lease "
         "linkage)   |   Read this page once, then go to the 'SD Register' tab.",
      size=10, color="6B7280", height=20)
s1.row_dimensions[3].height = 8

s1row(4, "How to fill this template", size=14, bold=True, color=BRAND, height=26)

steps = [
    ("Step 1", "Open the 'SD Register' tab. Your data goes in rows 7 to 26 — "
               "that is 20 deposits per file."),
    ("Step 2", "Rows 5 and 6 are worked SAMPLES, shown in grey italics. "
               "Delete both rows before you upload."),
    ("Step 3", "Click any cell to see a yellow help box telling you exactly what to "
               "enter. Hover over any orange or blue heading for a fuller note."),
    ("Step 4", "Orange headings are REQUIRED: Deposit Name, Deposit Amount, "
               "Market Rate, Tenure. Blue headings are optional — leave blank "
               "to use the default."),
    ("Step 5", "Columns with a small arrow on the right are dropdowns. Always pick "
               "from the list rather than typing, so the upload reads them correctly."),
    ("Step 6", "Save the file, then upload it in Finosutra → Security Deposit "
               "Portfolio → Upload Deposits. Check the preview screen before "
               "you confirm."),
]
r = 5
for label, body in steps:
    s1.row_dimensions[r].height = 32
    a = s1.cell(row=r, column=1, value=label.replace("Step ", ""))
    a.font = Font(name=F, size=11, bold=True, color=WHITE)
    a.fill = PatternFill("solid", fgColor=BRAND2)
    a.alignment = Alignment(horizontal="center", vertical="center")
    c = s1.cell(row=r, column=2, value=body)
    c.font = Font(name=F, size=11, color="111827")
    c.alignment = Alignment(vertical="center", wrap_text=True)
    r += 1

r += 1
s1row(r, "The five things people get wrong", size=14, bold=True, color=BRAND,
      height=28); r += 1
pitfalls = [
    "Tenure is in YEARS, not months. A deposit for 3 years 6 months is 3.5.",
    "Market Rate is a whole percentage. Enter 10 for 10% — not 0.10.",
    "Deposit Amount is the NOMINAL amount actually paid or received — not its "
    "present value. The present value and Day-1 discount are calculated for you.",
    "Deposit Type left blank defaults to 'Given' (you paid it). If your entity "
    "RECEIVED the deposit, you must type 'Received' — it does not default that way.",
    "Commencement Date is optional, but leaving it blank means the report shows "
    "lifetime totals only — it cannot split interest by financial year.",
]
for p in pitfalls:
    s1.row_dimensions[r].height = 30
    a = s1.cell(row=r, column=1, value="!")
    a.font = Font(name=F, size=12, bold=True, color=STOPFG)
    a.alignment = Alignment(horizontal="center", vertical="center")
    c = s1.cell(row=r, column=2, value=p)
    c.font = Font(name=F, size=11, color="111827")
    c.alignment = Alignment(vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor=NOTEBG)
    r += 1

r += 1
s1.row_dimensions[r].height = 46
c = s1.cell(row=r, column=2,
            value="LIMIT — 20 deposits per file.\nThis keeps the upload fast and "
                  "lets you actually check the preview screen. If you have more "
                  "than 20 deposits, save a second copy of this file and upload "
                  "it separately.")
c.font = Font(name=F, size=11, bold=True, color=STOPFG)
c.fill = PatternFill("solid", fgColor=STOPBG)
c.alignment = Alignment(vertical="center", wrap_text=True)
s1.cell(row=r, column=1).fill = PatternFill("solid", fgColor=STOPBG)
r += 2

s1row(r, "Tabs in this workbook", size=14, bold=True, color=BRAND, height=26); r += 1
for t, d in [("SD Register", "Where you enter your deposits. This is the only tab we read."),
             ("Field Guide", "Every column explained, with the common mistake for each."),
             ("Examples", "Five fully worked deposits — Given, Received, fractional "
                          "tenure and IFRS.")]:
    s1.row_dimensions[r].height = 20
    c = s1.cell(row=r, column=2, value="%-14s  %s" % (t, d))
    c.font = Font(name=F, size=11, color="374151")
    c.alignment = Alignment(vertical="center")
    r += 1

# ═════════════════════════════════════════════════════════════════════════
# SHEET 2 — SD Register
# ═════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("SD Register")
ws.sheet_view.showGridLines = False

ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
t = ws.cell(row=1, column=1, value="Finosutra   |   IND AS 109 Security Deposit Register")
t.font = Font(name=F, size=16, bold=True, color=WHITE)
t.fill = PatternFill("solid", fgColor=BRAND)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 30

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOL)
t = ws.cell(row=2, column=1,
            value="Version 1.0   |   Maximum 20 deposits per file   |   "
                  "Enter your data in rows 7-26   |   Click any cell for help")
t.font = Font(name=F, size=10, color="C7D2FE")
t.fill = PatternFill("solid", fgColor=BRAND)
t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[2].height = 18

ws.row_dimensions[3].height = 20
ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
lg = ws.cell(row=3, column=1, value="ORANGE heading = required     BLUE heading = optional")
lg.font = Font(name=F, size=9, bold=True, color="374151")
lg.fill = PatternFill("solid", fgColor=GREY_BG)
lg.alignment = Alignment(horizontal="left", vertical="center", indent=1)

ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=7)
lg2 = ws.cell(row=3, column=5,
              value="Rows 5-6 are samples - delete them before uploading.")
lg2.font = Font(name=F, size=9, italic=True, color="6B7280")
lg2.fill = PatternFill("solid", fgColor=GREY_BG)
lg2.alignment = Alignment(horizontal="left", vertical="center")

ws.merge_cells(start_row=3, start_column=8, end_row=3, end_column=NCOL)
cnt = ws.cell(row=3, column=8,
              value="Enter up to %d deposits  ->  rows %d to %d"
                    % (MAX_DEPOSITS, DATA_START, DATA_END))
cnt.font = Font(name=F, size=9, bold=True, color=BRAND2)
cnt.fill = PatternFill("solid", fgColor=GREY_BG)
cnt.alignment = Alignment(horizontal="right", vertical="center", indent=1)

ws.row_dimensions[HDR_ROW].height = 42
for i, col in enumerate(COLUMNS, start=1):
    c = ws.cell(row=HDR_ROW, column=i, value=col["h"])
    c.font = Font(name=F, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=AMBER if col["req"] else BRAND2)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    note = col["tip"]
    if col["miss"] != "-":
        note += "\n\nCommon mistake: " + col["miss"]
    cm = Comment(note, "Finosutra")
    cm.width, cm.height = 300, 190
    c.comment = cm
    ws.column_dimensions[get_column_letter(i)].width = col["w"]

samples = [
    ["SAMPLE 1 - delete this row", "Raheja Corp", "ABC Pvt. Ltd.", "Office Space",
     500000, 10, 3, "01/04/2023", "Given", "IND AS"],
    ["SAMPLE 2 - delete this row", "Sunrise Logistics", "ABC Pvt. Ltd.", "Warehouse",
     300000, 9, 5, "01/07/2022", "Received", "IND AS"],
]
for ri, row in enumerate(samples):
    rr = SAMPLE_1 + ri
    ws.row_dimensions[rr].height = 20
    for ci, v in enumerate(row, start=1):
        c = ws.cell(row=rr, column=ci, value=v)
        c.font = Font(name=F, size=10, italic=True, color="6B7280")
        c.fill = PatternFill("solid", fgColor=GREY_BG)
        c.border = box
        c.number_format = COLUMNS[ci - 1]["fmt"]
        c.alignment = Alignment(horizontal="left" if COLUMNS[ci - 1]["fmt"] == "@"
                                else "right", vertical="center")

for rr in range(DATA_START, DATA_END + 1):
    ws.row_dimensions[rr].height = 19
    band = ALT_BG if (rr - DATA_START) % 2 else WHITE
    for ci, col in enumerate(COLUMNS, start=1):
        c = ws.cell(row=rr, column=ci)
        c.font = Font(name=F, size=10, color="111827")
        c.fill = PatternFill("solid", fgColor=band)
        c.border = box
        c.number_format = col["fmt"]
        c.alignment = Alignment(horizontal="left" if col["fmt"] == "@" else "right",
                                vertical="center")

ws.row_dimensions[STOP_ROW].height = 34
ws.merge_cells(start_row=STOP_ROW, start_column=1, end_row=STOP_ROW, end_column=NCOL)
sc = ws.cell(row=STOP_ROW, column=1,
             value="STOP  —  Maximum 20 deposits per file. Do not add rows below "
                   "this line. If you have more than 20 deposits, save a second "
                   "copy of this file and upload it separately.")
sc.font = Font(name=F, size=11, bold=True, color=STOPFG)
sc.fill = PatternFill("solid", fgColor=STOPBG)
sc.alignment = Alignment(horizontal="center", vertical="center")

DV_RANGE_END = 200

def add_dv(col_idx, formula1=None):
    col = COLUMNS[col_idx - 1]
    letter = get_column_letter(col_idx)
    if formula1:
        dv = DataValidation(type="list", formula1=formula1, allow_blank=True,
                            showDropDown=False, showErrorMessage=True,
                            errorTitle="Please pick from the list",
                            error="Use the dropdown arrow on the right of this cell. "
                                  "Typing your own text here can make the upload "
                                  "misread this deposit.",
                            errorStyle="warning")
    else:
        dv = DataValidation(allow_blank=True, showErrorMessage=False)
    dv.promptTitle = col["tt"][:32]
    dv.prompt = col["tip"][:255]
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add("%s%d:%s%d" % (letter, SAMPLE_1, letter, DV_RANGE_END))

DROPDOWNS = {
    4: "Lists!$A$2:$A$11",   # Asset / Property Category
    9: DD_TYPE,              # Deposit Type
    10: DD_STD,              # Reporting Standard
}
for i in range(1, NCOL + 1):
    add_dv(i, DROPDOWNS.get(i))

ws.freeze_panes = "A5"
ws.sheet_properties.tabColor = BRAND2

# ═════════════════════════════════════════════════════════════════════════
# SHEET 3 — Field Guide
# ═════════════════════════════════════════════════════════════════════════
fg = wb.create_sheet("Field Guide")
fg.sheet_view.showGridLines = False
widths = [7, 26, 11, 56, 34, 42]
for i, w in enumerate(widths, start=1):
    fg.column_dimensions[get_column_letter(i)].width = w

fg.merge_cells("A1:F1")
c = fg.cell(row=1, column=1, value="Field Guide  —  every column explained")
c.font = Font(name=F, size=15, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=BRAND)
c.alignment = Alignment(vertical="center", indent=1)
fg.row_dimensions[1].height = 30

hdrs = ["Col", "Field", "Required?", "What to enter", "Valid values / example",
        "Common mistake"]
fg.row_dimensions[2].height = 24
for i, h in enumerate(hdrs, start=1):
    c = fg.cell(row=2, column=i, value=h)
    c.font = Font(name=F, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=BRAND2)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = box

VALID = {
    1: "Any text. Must be unique.",
    2: "Any text.",
    3: "Any text.",
    4: "Dropdown: Office Space, Warehouse, Retail Store, Plant & Machinery, "
       "Vehicle, IT Equipment, Equipment, Land, Building, Other",
    5: "Number only.  500000",
    6: "Number.  10 = 10%,  10.5 = 10.5%",
    7: "Number.  3.5 = three and a half years",
    8: "DD/MM/YYYY   e.g. 01/04/2025",
    9: "Dropdown: Given / Received",
    10: "Dropdown: IND AS / IFRS",
}
r = 3
for i, col in enumerate(COLUMNS, start=1):
    fg.row_dimensions[r].height = 42
    vals = [get_column_letter(i), col["h"].replace(" *", ""),
            "REQUIRED" if col["req"] else "Optional",
            col["tip"].replace("\n", "  "), VALID[i], col["miss"]]
    for ci, v in enumerate(vals, start=1):
        c = fg.cell(row=r, column=ci, value=v)
        c.font = Font(name=F, size=9,
                      bold=(ci == 3 and col["req"]),
                      color=AMBER if (ci == 3 and col["req"]) else "111827")
        c.alignment = Alignment(vertical="top", wrap_text=True,
                                horizontal="center" if ci in (1, 3) else "left")
        c.border = box
        c.fill = PatternFill("solid", fgColor=WHITE if r % 2 else ALT_BG)
    r += 1

r += 1
fg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
c = fg.cell(row=r, column=1, value="Notes")
c.font = Font(name=F, size=12, bold=True, color=BRAND)
r += 1
notes = [
    "Deposit Amount is always the cash amount that changed hands - the present "
    "value (fair value) and Day-1 discount are calculated automatically from "
    "Market Rate and Tenure.",
    "A deposit with no Commencement Date is still valid - the amortisation "
    "schedule is built Year 1, Year 2... from Tenure alone, but the portfolio "
    "KPIs cannot show a single financial year's figures for it.",
    "Re-uploading a deposit with a name that already exists UPDATES it rather "
    "than creating a duplicate. Keep your deposit names stable between uploads.",
    "Do not rename, reorder, or delete the column headings in row 4 of the SD "
    "Register — the upload matches on those exact headings.",
]
for n in notes:
    fg.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    fg.row_dimensions[r].height = 30
    c = fg.cell(row=r, column=1, value="•  " + n)
    c.font = Font(name=F, size=10, color="374151")
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    r += 1

# ═════════════════════════════════════════════════════════════════════════
# SHEET 4 — Examples
# ═════════════════════════════════════════════════════════════════════════
ex = wb.create_sheet("Examples")
ex.sheet_view.showGridLines = False
ex.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
c = ex.cell(row=1, column=1,
            value="Worked examples  —  copy the pattern that matches your deposit")
c.font = Font(name=F, size=15, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=BRAND)
c.alignment = Alignment(vertical="center", indent=1)
ex.row_dimensions[1].height = 30

EXAMPLES = [
    ("A. Simple Given deposit — tenant pays landlord, flat tenure",
     ["Mumbai Office Deposit", "Raheja Corp", "ABC Pvt. Ltd.", "Office Space",
      500000, 10, 3, "01/04/2025", "Given", "IND AS"]),
    ("B. Received deposit — landlord holding a tenant's deposit",
     ["Pune Warehouse Deposit", "Sunrise Logistics", "ABC Pvt. Ltd.", "Warehouse",
      300000, 9, 5, "01/07/2022", "Received", "IND AS"]),
    ("C. Fractional tenure — 3 years 6 months, note Tenure = 3.5",
     ["Chennai Retail Deposit", "TN Malls Pvt Ltd", "XYZ Ltd.", "Retail Store",
      750000, 11, 3.5, "01/01/2024", "Given", "IND AS"]),
    ("D. IFRS reporting variant — same mechanics, IFRS 9 labels",
     ["Singapore Office Deposit", "Raffles Estates", "XYZ Global Pte Ltd", "Office Space",
      1200000, 6, 4, "01/06/2023", "Given", "IFRS"]),
    ("E. No Commencement Date — still valid, lifetime totals only",
     ["Legacy Godown Deposit", "Local Owner", "ABC Pvt. Ltd.", "Warehouse",
      150000, 8, 5, "", "Given", "IND AS"]),
]

r = 3
for i, col in enumerate(COLUMNS, start=1):
    c = ex.cell(row=r, column=i, value=col["h"])
    c.font = Font(name=F, size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=AMBER if col["req"] else BRAND2)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = box
    ex.column_dimensions[get_column_letter(i)].width = col["w"]
ex.row_dimensions[r].height = 40
r += 1

for title, row in EXAMPLES:
    ex.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOL)
    c = ex.cell(row=r, column=1, value=title)
    c.font = Font(name=F, size=10, bold=True, color=BRAND)
    c.fill = PatternFill("solid", fgColor=OKBG)
    c.alignment = Alignment(vertical="center", indent=1)
    ex.row_dimensions[r].height = 22
    r += 1
    for ci, v in enumerate(row, start=1):
        val = v if v != "" else None
        c = ex.cell(row=r, column=ci, value=val)
        c.font = Font(name=F, size=10, color="111827")
        c.number_format = COLUMNS[ci - 1]["fmt"]
        c.border = box
        c.alignment = Alignment(horizontal="left" if COLUMNS[ci - 1]["fmt"] == "@"
                                else "right", vertical="center")
    ex.row_dimensions[r].height = 19
    r += 2

ex.freeze_panes = "A4"

# ═════════════════════════════════════════════════════════════════════════
# SHEET 5 — Lists (hidden, drives the Asset / Property Category dropdown)
# ═════════════════════════════════════════════════════════════════════════
lst = wb.create_sheet("Lists")
lst.cell(row=1, column=1, value="Asset / Property Category").font = Font(name=F, bold=True)
for i, v in enumerate(CATEGORIES, start=2):
    lst.cell(row=i, column=1, value=v).font = Font(name=F, size=10)
lst.column_dimensions["A"].width = 26
lst.sheet_state = "hidden"

wb.active = 0
# Written to the repo root, where downloadTemplate() in sd-portfolio.html serves it from.
# Keep SD_TEMPLATE_URL in sd-portfolio.html in step with this filename.
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                   "Finosutra_SD_Template_v1.xlsx")
wb.save(OUT)
print("saved:", OUT)
